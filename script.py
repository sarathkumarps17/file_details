import os
from preview_generator.manager import PreviewManager
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import datetime
wb = Workbook()

imageFiles = []
docFiles = []
videoFiles = []

cache_path = './preview_cache'
manager = PreviewManager(cache_path, create_folder=True)


# create sheet for each type of files

def getMeta(file, fileType):
    return {
        'fileName': file,
        'fileType': fileType,
        'size': datetime.datetime.fromtimestamp(os.path.getsize(file)),
        'modified': datetime.datetime.fromtimestamp(os.path.getmtime(file)),
        'created': datetime.datetime.fromtimestamp(os.path.getctime(file))
    }


# create sheet for each type of files
def createExcelFile(worker, row_lenth):
    worker['A1'] = 'File Name'
    worker['B1'] = 'File Type'
    worker['C1'] = 'File Size'
    worker['D1'] = 'Created Date'
    worker['E1'] = "Modified Date"
    worker['F1'] = "Preview"
    for i in range(1, row_lenth+1):
        worker.row_dimensions[i+1].height = 200

# Add meta data for each type of files


def addMetaToSheet(pathArr, sheet):
    createExcelFile(sheet, len(pathArr))
    index = 2
    for fileData in pathArr:
        fileName = fileData["fileName"]
        preview_image = manager.get_jpeg_preview(fileName)
        sheet[f'A{index}'] = fileName
        sheet[f'B{index}'] = fileData["fileType"]
        sheet[f'C{index}'] = fileData["size"]
        sheet[f'D{index}'] = fileData["created"]
        sheet[f'E{index}'] = fileData["modified"]
        image = Image(preview_image)
        sheet.add_image(image, f'F{index}')
        # ws[f'F{index}'] = preview_image
        index += 1


# return all files as a list
for file in os.listdir('./'):
    # check the files which are end with specific extension
    if file.endswith((".png", ".jpg", ".jpeg", ".gif")):
        imageFiles.append(getMeta(file, 'image'))
    elif file.endswith(('pdf', '.doc', '.docx')):
        docFiles.append(getMeta(file, 'document'))
    elif file.endswith((".mpg", ".mp2", ".mpeg", ".mpe", ".mpv", ".mp4", ".m4p",
                        " .m4v", ".webm", "mkv", '.mov', '.flv')):
        videoFiles.append(getMeta(file, 'video'))


im_sheet = wb.create_sheet('image_files', 0)
doc_sheet = wb.create_sheet('doc_files', 1)
video_sheet = wb.create_sheet('video_files', 2)
addMetaToSheet(imageFiles, im_sheet)
addMetaToSheet(docFiles, doc_sheet)
addMetaToSheet(videoFiles, video_sheet)

wb.save('files_data.xlsx')
